# -*- coding: utf-8 -*-
"""Extrai o historico da oficina das planilhas originais + planilha mestra.
Saida: historico.json (dados normalizados) + relatorio no stdout.
NAO toca no banco."""
import openpyxl, sys, os, re, json, datetime, unicodedata
from collections import Counter, defaultdict
sys.stdout.reconfigure(encoding="utf-8")
import warnings; warnings.filterwarnings("ignore")

RAIZ = r"C:\Users\henri\Downloads\ARQUIVOS PRO HENRIQUE FUGA\ARQUIVOS PRO HENRIQUE FUGA\Arquivos Importantes"
P01 = os.path.join(RAIZ, "01 - Clientes e Cadastros")
MESTRA = os.path.join(RAIZ, "Planilha Mestra - Base para ERP.xlsx")

# Ordem = prioridade de canonicidade (o 1o vence em aba repetida).
# PALIO WEKEEND e o save mais recente e superset de CROSSFOX e CORSA DANIEL.
FONTES = ["PALIO WEKEEND.xlsx", "CROSSFOX.xlsx", "CORSA DANIEL.xlsx",
          "SERVIÇO.xlsx", "SEMANA 4.xlsx",
          "PLANILHA DE CLIENTES ANTIGOS 02.xlsx", "PLANILHA DE CLIENTES ANTIGOS.xlsx"]

# ---------------------------------------------------------------- utilidades
def sem_acento(s):
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()

def norm_nome(s):
    """Normalizacao para dedup de cliente: acento/caixa/espaco APENAS."""
    if s is None: return ""
    t = sem_acento(s).upper()
    t = re.sub(r"[^A-Z0-9 ]", " ", t)
    return re.sub(r"\s+", " ", t).strip()

def norm_aba(s):
    t = sem_acento(s).upper().replace("_X0009_", " ")
    return re.sub(r"[^A-Z0-9]", "", t)

def norm_placa(s):
    if s is None: return None
    t = re.sub(r"[^A-Z0-9]", "", sem_acento(s).upper())
    if re.fullmatch(r"[A-Z]{3}\d[A-Z]\d{2}", t) or re.fullmatch(r"[A-Z]{3}\d{4}", t):
        return t
    return None

def num(v):
    if v is None: return None
    if isinstance(v, bool): return None
    if isinstance(v, (int, float)): return float(v)
    t = str(v).strip()
    if not t: return None
    t = re.sub(r"[R$\s]", "", t)
    if "," in t and "." in t: t = t.replace(".", "").replace(",", ".")
    elif "," in t: t = t.replace(",", ".")
    try:
        f = float(t)
        return f if f == f and abs(f) != float("inf") else None
    except ValueError:
        return None

def txt(c):
    return sem_acento(c).strip().upper() if c is not None else ""

def data_iso(v):
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    if isinstance(v, str):
        m = re.search(r"\b(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})\b", v)
        if m:
            d, mo, a = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if a < 100: a += 2000
            try: return datetime.date(a, mo, d).isoformat()
            except ValueError: return None
    return None

# ------------------------------------------------- parsing de bloco (aba OS)
ROTULOS = ("CHAVE PIX", "NOME:", "DATA", "CLIENTE", "VEICULO", "PLACA",
           "MAO DE OBRA", "VALOR TOTAL", "SERVICO FEITO POR", "QTD",
           "ITEM N", "DESCRICAO", "PRECO UNITARIO", "TOTAL", "MECANICA ALTAS HORAS")

def valor_apos(linha, j):
    for k in range(j + 1, len(linha)):
        if linha[k] is not None and str(linha[k]).strip():
            return linha[k]
    return None

def acha_rotulo(bloco, prefixo):
    for i, r in enumerate(bloco):
        for j, c in enumerate(r):
            if txt(c).startswith(prefixo):
                return i, j
    return None, None

def parse_bloco(bloco, origem):
    reg = {"origem": origem, "data": None, "cliente": None, "veiculo": None,
           "placa_raw": None, "placa": None, "tecnico": None, "itens": [],
           "mao_de_obra": None, "valor_total": None}

    for campo, pref in (("data", "DATA"), ("cliente", "CLIENTE:"),
                        ("veiculo", "VEICULO:"), ("placa_raw", "PLACA/KM:")):
        i, j = acha_rotulo(bloco, pref)
        if i is not None:
            v = valor_apos(bloco[i], j)
            if campo == "data": reg["data"] = data_iso(v)
            elif v is not None: reg[campo] = str(v).strip()
    reg["placa"] = norm_placa(reg["placa_raw"])

    i, _ = acha_rotulo(bloco, "SERVICO FEITO POR")
    if i is not None and i + 1 < len(bloco):
        for c in bloco[i + 1]:
            if c is not None and str(c).strip() and not txt(c).startswith(ROTULOS):
                reg["tecnico"] = str(c).strip(); break

    # cabecalho da tabela de itens -> mapeia colunas
    hi = hj = None
    for i, r in enumerate(bloco):
        for j, c in enumerate(r):
            if txt(c) == "QTD": hi, hj = i, j; break
        if hi is not None: break

    if hi is not None:
        hdr = bloco[hi]
        col = {}
        for j, c in enumerate(hdr):
            t = txt(c)
            if t == "QTD": col["qtd"] = j
            elif t.startswith("DESCRICAO"): col["desc"] = j
            elif t.startswith("PRECO"): col["pu"] = j
            elif t == "TOTAL": col["tot"] = j
        for i in range(hi + 1, len(bloco)):
            r = bloco[i]
            if any(txt(c).startswith(("MAO DE OBRA", "VALOR TOTAL", "CHAVE PIX")) for c in r):
                break
            desc = r[col["desc"]] if "desc" in col and col["desc"] < len(r) else None
            if desc is None or not str(desc).strip():
                cands = [c for jj, c in enumerate(r) if isinstance(c, str)
                         and len(c.strip()) > 2 and not txt(c).startswith(ROTULOS)]
                desc = cands[0] if cands else None
            if desc is None or not str(desc).strip(): continue
            if txt(desc).startswith(ROTULOS): continue
            q = num(r[col["qtd"]]) if "qtd" in col and col["qtd"] < len(r) else None
            pu = num(r[col["pu"]]) if "pu" in col and col["pu"] < len(r) else None
            tt = num(r[col["tot"]]) if "tot" in col and col["tot"] < len(r) else None
            if tt is None and pu is None:
                nums = [num(c) for c in r]
                nums = [n for n in nums if n is not None]
                if nums: tt = nums[-1]
            if tt is None and pu is None: continue
            reg["itens"].append({"descricao": str(desc).strip(),
                                 "quantidade": q if q and q > 0 else 1.0,
                                 "preco_unitario": pu, "total": tt})

    i, j = acha_rotulo(bloco, "MAO DE OBRA")
    if i is not None:
        v = valor_apos(bloco[i], j)
        reg["mao_de_obra"] = num(v)

    i, j = acha_rotulo(bloco, "VALOR TOTAL")
    if i is not None:
        v = num(valor_apos(bloco[i], j))
        if v is None and i + 1 < len(bloco):
            nums = [num(c) for c in bloco[i + 1]]
            nums = [n for n in nums if n is not None and n != 64996488838]
            v = nums[-1] if nums else None
        reg["valor_total"] = v
    return reg

def blocos_da_aba(ws):
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    starts = [i for i, r in enumerate(grid)
              if any(txt(c).startswith("SERVICO FEITO POR") for c in r)]
    if not starts:
        starts = [i for i, r in enumerate(grid)
                  if any(txt(c).startswith("CLIENTE:") for c in r)]
        starts = [max(0, i - 4) for i in starts]
    if not starts: return []
    out = []
    for k, s in enumerate(starts):
        e = starts[k + 1] if k + 1 < len(starts) else len(grid)
        out.append(grid[s:e])
    return out

# ------------------------------------------------------ 1) planilhas fonte
# DEDUP POR CONTEUDO, nao por nome de aba: dentro do MESMO arquivo existem abas
# que so diferem por espaco no nome ('CROSSFOX', 'CROSSFOX ', 'CROSSFOX  ') e
# que sao OS DIFERENTES (clientes/datas distintos). Deduplicar por nome apagaria
# registro real. A copia entre arquivos e pega pela impressao digital do bloco.
print("=" * 78); print("1) EXTRACAO DAS PLANILHAS ORIGINAIS (dedup por conteudo)")

def digital(r):
    itens = tuple(sorted((norm_nome(i["descricao"]), i["quantidade"], i["total"])
                         for i in r["itens"]))
    return (r["data"], norm_nome(r["cliente"]), norm_nome(r["veiculo"]),
            r["placa"], r["valor_total"], r["mao_de_obra"], itens)

registros, vistos, dup_conteudo = [], {}, []
for fn in FONTES:
    p = os.path.join(P01, fn)
    if not os.path.exists(p):
        print(f"  !! ausente: {fn}"); continue
    wb = openpyxl.load_workbook(p, data_only=True)
    novas = repet = vazias = 0
    for sn in wb.sheetnames:
        if re.fullmatch(r"(MODELO|PLANILHA\d+)\s*", sem_acento(sn).upper()): continue
        bl = blocos_da_aba(wb[sn])
        if not bl:
            vazias += 1; continue
        for i, b in enumerate(bl):
            rot = f"{fn} :: aba '{sn}'" + (f" [bloco {i+1}]" if len(bl) > 1 else "")
            r = parse_bloco(b, rot)
            if not (r["cliente"] or r["veiculo"] or r["itens"]):
                vazias += 1; continue
            d = digital(r)
            if d in vistos:
                repet += 1; dup_conteudo.append((rot, vistos[d])); continue
            vistos[d] = rot
            registros.append(r); novas += 1
    print(f"  {fn:38} OS novas={novas:4}  copias descartadas={repet:4}  vazias={vazias:3}")

print(f"\n  OS unicas extraidas das planilhas: {len(registros)}")
print(f"  blocos descartados por serem copia identica: {len(dup_conteudo)}")
print("  exemplos de copia descartada:")
for novo, orig in dup_conteudo[:5]:
    print(f"     {novo}\n        == {orig}")

# ------------------------------------------------------ 2) planilha mestra
print("\n" + "=" * 78); print("2) PLANILHA MESTRA (PDF/DOCX das pastas 02 e 03)")
wbm = openpyxl.load_workbook(MESTRA, data_only=True)
def linhas(aba):
    d = list(wbm[aba].iter_rows(values_only=True))
    h = [str(x) if x is not None else "" for x in d[0]]
    return [dict(zip(h, r)) for r in d[1:] if any(c is not None for c in r)]

MO = linhas("Ordens_e_Orcamentos")
docs = [o for o in MO if "planilha" not in str(o["Tipo"])]
seen, docs_u = set(), []
for o in docs:
    k = (norm_nome(o.get("Cliente")), norm_nome(o.get("Veiculo")),
         norm_placa(o.get("Placa_KM")), o.get("Data"), num(o.get("Valor_Total")))
    if k in seen: continue
    seen.add(k); docs_u.append(o)
print(f"  linhas PDF/DOCX: {len(docs)} -> unicas: {len(docs_u)}")
print(f"    Ordem de Servico: {sum(1 for o in docs_u if o['Tipo']=='Ordem de Servico')}")
print(f"    Orcamento:        {sum(1 for o in docs_u if o['Tipo']=='Orcamento')}")

docs_reg = []
for o in docs_u:
    docs_reg.append({
        "origem": f"planilha-mestra :: {o['Arquivo_Fonte']}",
        "arquivo": str(o["Arquivo_Fonte"]),
        "pasta": str(o["Origem"]),
        "tipo": "orcamento" if o["Tipo"] == "Orcamento" else "os",
        "data": data_iso(o.get("Data")), "cliente": o.get("Cliente"),
        "veiculo": o.get("Veiculo"), "placa": norm_placa(o.get("Placa_KM")),
        "placa_raw": o.get("Placa_KM"), "tecnico": None, "itens": [],
        "mao_de_obra": None, "valor_total": num(o.get("Valor_Total"))})

for r in registros:
    r["tipo"] = "os"; r["arquivo"] = r["origem"].split(" :: ")[0]; r["pasta"] = "01 - Clientes e Cadastros"
todos = registros + docs_reg

# ------------------------------------------------------ 3) financeiro
print("\n" + "=" * 78); print("3) FINANCEIRO")
FIN = linhas("Transacoes_Financeiro")
MES = {"JANEIRO":1,"FEVEREIRO":2,"MARCO":3,"ABRIL":4,"MAIO":5,"JUNHO":6,"JULHO":7,
       "AGOSTO":8,"SETEMBRO":9,"OUTUBRO":10,"NOVEMBRO":11,"DEZEMBRO":12}
def forma_pgto(s):
    t = sem_acento(s or "").upper()
    if not t.strip(): return None, None
    obs = s.strip() if any(x in t for x in ("FALTA","RECEBEU","FIADO","/")) else None
    if "FIADO" in t: return None, obs
    if "PIX" in t and "CART" in t: return "pix", obs
    if "PIX" in t: return "pix", obs
    if "DEBITO" in t: return "cartao_debito", obs
    if "CREDITO" in t: return "cartao_credito", obs
    if "CART" in t: return "cartao_credito", obs
    if "DINHEIRO" in t: return "dinheiro", obs
    return None, obs

# Ancoras de data por arquivo-fonte, apuradas lendo os proprios arquivos em
# 04 - Financeiro (ver fin_fontes.py):
#   Planilha Diaria.xlsx ......... datas reais por linha (abr/2023)
#   PLANILHA BRUTO JULHO.xlsx .... sem data interna; titulo "Planilha JULHO",
#                                  arquivo modificado em 14/08/2023 -> jul/2023
#   PLANILHA MENSAL SETEMBRO.xlsx  titulo traz "11/09/2023" explicito
ANCORA = {"PLANILHA BRUTO JULHO.XLSX": "2023-07-01",
          "PLANILHA MENSAL SETEMBRO.XLSX": "2023-09-11"}

fin, sem_data_fin = [], 0
for i, f in enumerate(FIN, start=2):
    tot = num(f.get("TOTAL")); pcs = num(f.get("PEÇAS")); mo = num(f.get("MÃO DE OBRA"))
    d = data_iso(f.get("ENTRADA")) or data_iso(f.get("DATA DE ENTREGA"))
    fonte = str(f.get("_Fonte") or "")
    inferida = False
    if not d:
        chave = sem_acento(fonte).upper().strip()
        if chave in ANCORA:
            d = ANCORA[chave]; inferida = True
        else:
            mm = next((v for k, v in MES.items() if k in sem_acento(fonte).upper()), None)
            if mm: d = datetime.date(2023, mm, 1).isoformat(); inferida = True
            else: sem_data_fin += 1
    fp, obs = forma_pgto(f.get("DINHEIRO / CARTÃO") or f.get("DINHEIRO /CARTÃO"))
    fin.append({"linha": i, "cliente": f.get("CLIENTE"), "contato": f.get("CONTATO"),
                "veiculo": f.get("CARRO/Ano"),
                "servico": (f.get("SERVIÇO") or f.get("SERVIÇO FEITO")),
                "pecas": pcs, "mao_de_obra": mo, "total": tot, "data": d,
                "data_inferida": inferida, "forma_pagamento": fp, "obs": obs,
                "fonte": fonte, "origem": f"planilha-mestra :: Transacoes_Financeiro L{i}"})
com_tot = [x for x in fin if x["total"] and x["total"] > 0]
print(f"  linhas: {len(fin)}  com TOTAL>0: {len(com_tot)}  soma: R$ {sum(x['total'] for x in com_tot):,.2f}")
print(f"  com data real: {sum(1 for x in fin if x['data'] and not x['data_inferida'])}"
      f"  | data inferida do nome do arquivo: {sum(1 for x in fin if x['data_inferida'])}"
      f"  | sem data: {sem_data_fin}")
print(f"  forma de pagamento identificada: {sum(1 for x in fin if x['forma_pagamento'])}/{len(fin)}")
print("  distribuicao:", dict(Counter(x["forma_pagamento"] for x in fin)))

# ------------------------------------------------------ 4) entidades
print("\n" + "=" * 78); print("4) ENTIDADES (dedup por acento/caixa/espaco)")
CU = linhas("Clientes_Unicos")
MARCAS = set("""GOL UNO PALIO CORSA CELTA CIVIC FIESTA KA STRADA SAVEIRO VECTRA ASTRA MONTANA
S10 HILUX ONIX PRISMA VOYAGE FOX CROSSFOX SPIN TORO PUNTO IDEA SIENA LOGAN SANDERO DUSTER
ARGO MOBI CRONOS POLO VIRTUS JETTA GOLF COROLLA ETIOS HB20 CRETA TUCSON RANGER AMAROK FIAT
VW FORD CHEVROLET HONDA TOYOTA RENAULT NISSAN PEUGEOT CITROEN BRAVA TEMPRA MAREA PARATI
SANTANA KOMBI BLAZER CLASSIC AGILE COBALT TRACKER KICKS PULSE FASTBACK DOBLO CLIO FOCUS
FUSION LANCER JEEP BRAVO WEEKEND WEKEEND VIVACE G2 G3 G4 G5 G6 G7 G8""".split())

def parece_veiculo(nome):
    toks = norm_nome(nome).split()
    return bool(toks) and all(t in MARCAS or t.isdigit() for t in toks)

clientes = {}
def reg_cliente(nome, origem):
    k = norm_nome(nome)
    if not k: return None
    c = clientes.get(k)
    if not c:
        c = clientes[k] = {"chave": k, "nome": str(nome).strip(), "variantes": set(),
                           "origens": set(), "veiculo_like": parece_veiculo(nome)}
    c["variantes"].add(str(nome).strip()); c["origens"].add(origem)
    # prefere a grafia mais "cuidada" (com acento/maiuscula-minuscula mista)
    if len(str(nome).strip()) > len(c["nome"]): c["nome"] = str(nome).strip()
    return k

for c in CU: reg_cliente(c["Cliente"], "Clientes_Unicos")
for r in todos: reg_cliente(r.get("cliente"), r["origem"])
for f in fin: reg_cliente(f.get("cliente"), f["origem"])

print(f"  clientes unicos apos dedup: {len(clientes)}")
print(f"    (Clientes_Unicos tinha {len(CU)}; uniao com todas as fontes, normalizada)")
vl = [c for c in clientes.values() if c["veiculo_like"]]
print(f"  marcados 'identificacao pendente' (nome e carro): {len(vl)}")
print("    ex:", sorted(c["nome"] for c in vl)[:14])
mult = [c for c in clientes.values() if len({norm_nome(v) for v in c["variantes"]}) >= 1 and len(c["variantes"]) > 1]
print(f"  clientes que juntaram >1 grafia: {len(mult)}")
for c in sorted(mult, key=lambda x: -len(x["variantes"]))[:12]:
    print(f"     {c['nome']:26} <- {sorted(c['variantes'])}")

# veiculos: chave = placa quando existe, senao (cliente, modelo normalizado)
veiculos, sem_placa_n = {}, 0
def reg_veiculo(cli_k, modelo, placa, origem):
    global sem_placa_n
    if not modelo and not placa: return None
    if placa: k = ("placa", placa)
    else:
        if not cli_k: return None
        k = ("cli", cli_k, norm_nome(modelo))
    v = veiculos.get(k)
    if not v:
        v = veiculos[k] = {"chave": k, "placa": placa, "modelos": set(),
                           "clientes": set(), "origens": set()}
    if modelo: v["modelos"].add(str(modelo).strip())
    if cli_k: v["clientes"].add(cli_k)
    v["origens"].add(origem)
    return k

# 1a passada: so os que TEM placa, para saber (cliente, modelo) -> placa
placa_de = defaultdict(set)
for r in todos:
    if r.get("placa"):
        placa_de[(norm_nome(r.get("cliente")), norm_nome(r.get("veiculo")))].add(r["placa"])
# so vale como ancora quando ha UMA placa possivel (senao seria chute)
ancora = {k: next(iter(v)) for k, v in placa_de.items() if len(v) == 1}

herdadas = 0
for r in todos:
    p = r.get("placa")
    if not p:
        p = ancora.get((norm_nome(r.get("cliente")), norm_nome(r.get("veiculo"))))
        if p:
            herdadas += 1
            r["placa_herdada"] = True   # placa deduzida de outro registro do mesmo cliente+modelo
            r["placa"] = p
    reg_veiculo(norm_nome(r.get("cliente")), r.get("veiculo"), p, r["origem"])
print(f"  registros sem placa que herdaram a placa do mesmo cliente+modelo: {herdadas}")

for v in veiculos.values():
    if not v["placa"]:
        sem_placa_n += 1
        v["placa_sintetica"] = f"SEM-PLACA-{sem_placa_n:04d}"
print(f"\n  veiculos unicos: {len(veiculos)}  (com placa real: {len(veiculos)-sem_placa_n}"
      f" | placa sintetica: {sem_placa_n})")
compart = [v for v in veiculos.values() if len(v["clientes"]) > 1]
print(f"  !! veiculos com >1 cliente (mesma placa, nomes diferentes): {len(compart)}")
for v in compart[:10]:
    print(f"     {v['placa']}: {sorted(v['clientes'])}  modelos={sorted(v['modelos'])}")

# ------------------------------------------------------ 5) qualidade das OS
print("\n" + "=" * 78); print("5) QUALIDADE DAS OS / ORCAMENTOS")
os_regs = [r for r in todos if r["tipo"] == "os"]
orc_regs = [r for r in todos if r["tipo"] == "orcamento"]
print(f"  OS: {len(os_regs)}  |  Orcamentos: {len(orc_regs)}  |  TOTAL: {len(todos)}")
for nome, grupo in (("OS", os_regs), ("Orcamento", orc_regs), ("TODOS", todos)):
    d = sum(1 for r in grupo if r["data"]); p = sum(1 for r in grupo if r["placa"])
    v = sum(1 for r in grupo if r["valor_total"]); it = sum(len(r["itens"]) for r in grupo)
    t = sum(1 for r in grupo if r.get("tecnico"))
    n = max(len(grupo), 1)
    print(f"  {nome:10} data={d:3}({100*d//n:3}%) placa={p:3}({100*p//n:3}%) "
          f"valor={v:3}({100*v//n:3}%) tecnico={t:3} itens={it:5}")

itens_tot = sum(len(r["itens"]) for r in todos)
soma_itens = sum(i["total"] or 0 for r in todos for i in r["itens"])
print(f"\n  linhas de item: {itens_tot}  soma R$ {soma_itens:,.2f}")
print(f"  mao de obra informada: {sum(1 for r in todos if r['mao_de_obra'])} "
      f"(R$ {sum(r['mao_de_obra'] or 0 for r in todos):,.2f})")

# coerencia: valor_total declarado x soma dos itens + mao de obra
div = []
for r in todos:
    if r["valor_total"] and r["itens"]:
        calc = sum(i["total"] or 0 for i in r["itens"]) + (r["mao_de_obra"] or 0)
        if calc > 0 and abs(calc - r["valor_total"]) > 0.5:
            div.append((r["origem"], r["valor_total"], round(calc, 2)))
print(f"  !! valor_total != soma(itens)+mao de obra: {len(div)} de "
      f"{sum(1 for r in todos if r['valor_total'] and r['itens'])} comparaveis")
for o, d1, d2 in div[:10]: print(f"     declarado={d1:>9} calculado={d2:>9}  {o}")

anos = Counter(r["data"][:4] for r in todos if r["data"])
print(f"\n  distribuicao por ano: {dict(sorted(anos.items()))}")
print(f"  sem data nenhuma: {sum(1 for r in todos if not r['data'])}")

# ------------------------------------------------------ 6) dump
saida = {
    "gerado_em": datetime.datetime.now().isoformat(timespec="seconds"),
    "clientes": [{**c, "variantes": sorted(c["variantes"]), "origens": sorted(c["origens"])[:3]}
                 for c in clientes.values()],
    "veiculos": [{**v, "chave": list(v["chave"]), "modelos": sorted(v["modelos"]),
                  "clientes": sorted(v["clientes"]), "origens": sorted(v["origens"])[:3]}
                 for v in veiculos.values()],
    "ordens": todos,
    "financeiro": fin,
}
dest = os.path.join(os.path.dirname(os.path.abspath(__file__)), "historico.json")
with open(dest, "w", encoding="utf-8") as fh:
    json.dump(saida, fh, ensure_ascii=False, indent=1)
print(f"\n  -> {dest}  ({os.path.getsize(dest)/1024:.0f} KB)")
